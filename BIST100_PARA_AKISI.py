import os
import warnings
import numpy as np
import pandas as pd
import requests
import yfinance as yf
import matplotlib.pyplot as plt
import io
from io import StringIO
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.drawing.image import Image
from openpyxl.chart import PieChart, Reference 

# =========================================================
# AYARLAR VE SEKTÖR ÇEVİRİ SÖZLÜĞÜ (Mapping Layer)
# =========================================================
warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", category=UserWarning)

SECTOR_TR_MAP = {
    "Financial Services": "Finansal Hizmetler",
    "Industrials": "Sanayi",
    "Technology": "Teknoloji",
    "Basic Materials": "Temel Maddeler",
    "Consumer Cyclical": "Tüketici Döngüsel",
    "Consumer Defensive": "Tüketici Savunma",
    "Healthcare": "Sağlık",
    "Utilities": "Kamu Hizmetleri",
    "Real Estate": "Gayrimenkul",
    "Communication Services": "İletişim Hizmetleri",
    "Energy": "Enerji",
    "Bilinmiyor": "Bilinmiyor"
}

LOOKBACK_DAYS = 10
CHART_DAYS = 20 
REG_PERIOD = 89  
DOWNLOAD_PERIOD = "7mo" 
OYAK_URL = "https://www.oyakyatirim.com.tr/piyasa-verileri/XU100"
OUTPUT_DIR = "/Users/yusufemreozden/Desktop/"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "bist100_para_akisi_final.xlsx")

SHEET_MAIN = "BIST 100 Piyasa Analizi"
SHEET_TOP = "Skor Bazlı Öncü Hisseler"
SHEET_SECTOR = "Sektörel Para Girişi" 
SHEET_SYMBOLS = "Çekilen BIST100 Listesi"
SHEET_SKIPPED = "Hata ve Eksik Veri Günlüğü"

COLUMN_MAPPING = {
    "symbol": "Hisse Kodu",
    "signed_flow_10d": "Net Para Akışı (10G)",
    "turnover_10d": "Toplam İşlem Hacmi (10G)",
    "flow_ratio_10d": "Para Akış Oranı",
    "positive_flow_10d": "Pozitif Akış Toplamı",
    "negative_flow_10d": "Negatif Akış Toplamı",
    "positive_days_10d": "Pozitif Gün Sayısı",
    "negative_days_10d": "Negatif Gün Sayısı",
    "return_10d_pct": "10 Günlük Getiri (%)",
    "avg_turnover_10d": "Ortalama Günlük Hacim",
    "reg_89": "Regresyon",
    "flow_strength_score": "Akış Gücü Skoru",
    "flow_comment": "Analiz Notu",
    "chart": "Grafik"
}

TL_FORMAT_COLS = ["Net Para Akışı (10G)", "Toplam İşlem Hacmi (10G)", "Pozitif Akış Toplamı", "Negatif Akış Toplamı", "Ortalama Günlük Hacim", "Toplam Net Para Akışı", "Sektörel Hacim"]
PCT_FORMAT_COLS = ["Para Akış Oranı", "10 Günlük Getiri (%)", "Hacim Ağırlıklı Getiri (%)"]
CONDITIONAL_COLS = ["Net Para Akışı (10G)", "Para Akış Oranı", "Akış Gücü Skoru", "10 Günlük Getiri (%)", "Toplam Net Para Akışı", "Ağırlıklı Akış Skoru"]
USER_AGENT = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}

# =========================================================
# MODÜL 1: VERİ ÇEKME
# =========================================================
def fetch_bist_data():
    r = requests.get(OYAK_URL, headers=USER_AGENT, timeout=20)
    r.raise_for_status()
    tables = pd.read_html(StringIO(r.text))
    
    target_tbl = None
    for t in tables:
        cols = [str(c).strip().lower() for c in t.columns]
        if "sembol" in " ".join(cols) and ("hacim" in " ".join(cols) or "son" in " ".join(cols)):
            target_tbl = t
            break
            
    if target_tbl is None: raise ValueError("XU100 tablosu bulunamadı.")
    
    symbol_col = [c for c in target_tbl.columns if "sembol" in str(c).lower()][0]
    raw_symbols = target_tbl[symbol_col].astype(str).str.strip().str.upper().dropna().tolist()
    clean_symbols = sorted(list(set([s for s in raw_symbols if s.isalpha() and 2 <= len(s) <= 6])))
    yahoo_symbols = [f"{s}.IS" for s in clean_symbols]
    
    data = yf.download(yahoo_symbols, period=DOWNLOAD_PERIOD, group_by="ticker", threads=True, progress=False, timeout=30, auto_adjust=True)
    
    sector_map = {}
    print(">>> Sektör bilgileri kurumsal dile çevriliyor...")
    for s in yahoo_symbols:
        try:
            info = yf.Ticker(s).info
            eng_sector = info.get('sector', 'Bilinmiyor')
            sector_map[s.replace(".IS", "")] = SECTOR_TR_MAP.get(eng_sector, eng_sector)
        except:
            sector_map[s.replace(".IS", "")] = 'Bilinmiyor'
            
    return data, clean_symbols, yahoo_symbols, sector_map

# =========================================================
# MODÜL 2: ANALİZ VE KANTİTATİF INSIGHT
# =========================================================
def safe_zscore(series):
    s = pd.to_numeric(series, errors="coerce").astype(float)
    std = s.std(ddof=0)
    return (s - s.mean()) / std if not (pd.isna(std) or std == 0) else pd.Series(np.zeros(len(s)), index=s.index)

def get_pro_insight(row):
    ratio_z, ret, score = row["ratio_zscore"], row["return_10d_pct"], row["flow_strength_score"]
    if ratio_z > 1.5 and ret > 5: return "Agresif Kurumsal Alım: Trend Hacimle Onaylanıyor"
    if ratio_z > 1.0 and -3 <= ret <= 4: return "Sessiz Biriktirme: Fiyat Baskılı, Para Girişi Güçlü"
    if ratio_z < -0.5 and ret > 7: return "Zayıf Yükseliş: Momentum Hacimle Desteklenmiyor (Riskli)"
    if ratio_z < -1.5 and ret < -5: return "Sert Dağıtım: Kurumsal Çıkış ve Satış Baskısı"
    if ratio_z > 1.0 and ret < -7: return "Tuzak/Destek Arayışı: Sert Düşüşe Karşı Kurumsal Alım"
    if score > 1.5: return "Pozitif Momentum: Güçlü İstatistiksel Ayrışma"
    if score < -1.5: return "Negatif Momentum: Güçlü İstatistiksel Satış"
    if ratio_z > 0.5: return "Pozitif Akış: Para Girişi Pozitif"
    if ratio_z < -0.5: return "Negatif Akış: Para Çıkışı Hissediliyor"
    return "Nötr: Belirgin Bir Akış Sapması Yok"

def create_sparkline_bytes(data_series):
    plt.figure(figsize=(2, 0.5))
    plt.plot(data_series.values, color='#2C3E50', linewidth=2)
    for spine in plt.gca().spines.values(): spine.set_visible(False)
    plt.xticks([]); plt.yticks([]); plt.grid(False)
    img_buf = io.BytesIO()
    plt.savefig(img_buf, format='png', bbox_inches='tight', pad_inches=0.05, transparent=True, dpi=100)
    plt.close()
    return img_buf.getvalue()

def perform_analysis(data, yahoo_symbols, sector_map):
    results, skipped, chart_data_bytes = [], [], {}
    for s in yahoo_symbols:
        symbol_clean = s.replace(".IS", "")
        try:
            if s not in data.columns.get_level_values(0) or data[s].empty:
                skipped.append({"Hisse": symbol_clean, "Neden": "Yahoo Finance verisi bulunamadı."}); continue
            
            df_full = data[s].copy().dropna(subset=["High", "Low", "Close", "Volume"])
            df_full = df_full[(df_full["Close"] > 0) & (df_full["Volume"] > 0)]
            df_flow = df_full.tail(LOOKBACK_DAYS + 1)
            
            if len(df_flow) < LOOKBACK_DAYS + 1:
                skipped.append({"Hisse": symbol_clean, "Neden": "Yetersiz veri günü"}); continue

            if len(df_full) >= REG_PERIOD:
                reg_prices = df_full["Close"].tail(REG_PERIOD).values
                x_axis = np.arange(len(reg_prices))
                slope, _ = np.polyfit(x_axis, reg_prices, 1)
                pearson_val = np.corrcoef(x_axis, reg_prices)[0, 1]
                if pearson_val >= 0.7 and slope > 0: reg_status = "Güçlü Pozitif"
                elif pearson_val <= -0.7 and slope < 0: reg_status = "Güçlü Negatif"
                else: reg_status = "Pozitif" if slope > 0 else "Negatif"
            else: reg_status = "Yetersiz Veri"

            chart_data_bytes[symbol_clean] = create_sparkline_bytes(df_full.tail(CHART_DAYS)["Close"])
            tp = (df_flow["High"] + df_flow["Low"] + df_flow["Close"]) / 3
            v = pd.DataFrame({"flow": tp * df_flow["Volume"], "sign": np.sign(tp - tp.shift(1))}).iloc[1:].copy()
            v["s_flow"] = v["flow"] * v["sign"]
            
            results.append({
                "symbol": symbol_clean, 
                "sector": sector_map.get(symbol_clean, 'Bilinmiyor'),
                "signed_flow_10d": v["s_flow"].sum(), 
                "recent_flow_3d": v["s_flow"].tail(3).sum(),
                "turnover_10d": v["flow"].sum(),
                "flow_ratio_10d": v["s_flow"].sum() / v["flow"].sum() if v["flow"].sum() != 0 else 0,
                "return_10d_pct": ((df_flow["Close"].iloc[-1] / df_flow["Close"].iloc[0]) - 1) * 100, 
                "avg_turnover_10d": v["flow"].mean(),
                "reg_89": reg_status,
                "chart": ""
            })
        except Exception as e:
            skipped.append({"Hisse": symbol_clean, "Neden": str(e)})
            
    res = pd.DataFrame(results)
    res["ratio_zscore"] = safe_zscore(res["flow_ratio_10d"])
    res["ret_zscore"] = safe_zscore(res["return_10d_pct"])
    res["flow_strength_score"] = (0.45 * res["ratio_zscore"] + 0.25 * safe_zscore(res["signed_flow_10d"]) + 0.30 * res["ret_zscore"])
    res["flow_comment"] = res.apply(get_pro_insight, axis=1)
    
    res['ret_weighted'] = res['return_10d_pct'] * res['turnover_10d']
    res['score_weighted'] = res['flow_strength_score'] * res['turnover_10d']
    
    sector_summary = res.groupby("sector").agg({
        "signed_flow_10d": "sum",
        "recent_flow_3d": "sum",
        "turnover_10d": "sum",
        "ret_weighted": "sum",
        "score_weighted": "sum",
        "symbol": "count"
    }).rename(columns={"signed_flow_10d": "Toplam Net Para Akışı", "turnover_10d": "Sektörel Hacim", "symbol": "Hisse Sayısı"})
    
    sector_summary['Hacim Ağırlıklı Getiri (%)'] = sector_summary['ret_weighted'] / sector_summary['Sektörel Hacim']
    sector_summary['Ağırlıklı Akış Skoru'] = sector_summary['score_weighted'] / sector_summary['Sektörel Hacim']
    
    sector_summary["Momentum"] = sector_summary.apply(
        lambda r: "Isınmaya Başladı" if r["Toplam Net Para Akışı"] > 0 and r["recent_flow_3d"] > (r["Toplam Net Para Akışı"] * 0.40) else 
                  ("Soğuyor" if r["Toplam Net Para Akışı"] > 0 and r["recent_flow_3d"] < 0 else "Dengeli"), axis=1)
    
    sector_summary = sector_summary.drop(columns=['ret_weighted', 'score_weighted', 'recent_flow_3d'])
    sector_summary = sector_summary.sort_values("Ağırlıklı Akış Skoru", ascending=False)
    
    # 1- SEKTÖR KOLON ADINI "Sektör" YAP
    sector_summary.index.name = "Sektör"

    return res.drop(columns=["ratio_zscore", "ret_zscore", "ret_weighted", "score_weighted", "recent_flow_3d"]).rename(columns=COLUMN_MAPPING), skipped, chart_data_bytes, sector_summary

# =========================================================
# MODÜL 3: FORMATLAMA VE GÖRSELLEŞTİRME
# =========================================================
def apply_excel_styling(ws):
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    zebra_fill = PatternFill("solid", fgColor="F8F9F9")
    fills = {
        "pos_strong": PatternFill("solid", fgColor="27AE60"), "pos_light": PatternFill("solid", fgColor="C6EFCE"),
        "neg_strong": PatternFill("solid", fgColor="C0392B"), "neg_light": PatternFill("solid", fgColor="FFC7CE"),
        "neutral": PatternFill("solid", fgColor="D5DBDB"), "heat": PatternFill("solid", fgColor="FAD7A0")
    }
    border = Border(left=Side(style='thin', color="D5D8DC"), right=Side(style='thin', color="D5D8DC"), top=Side(style='thin', color="D5D8DC"), bottom=Side(style='thin', color="D5D8DC"))
    
    for cell in ws[1]:
        cell.fill, cell.font, cell.border = header_fill, header_font, border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        if ws.title in [SHEET_MAIN, SHEET_TOP]: ws.row_dimensions[row_idx].height = 40 
        for cell in row:
            cell.border = border
            if row_idx % 2 == 0: cell.fill = zebra_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            col_name = ws.cell(row=1, column=cell.column).value
            if col_name == "Momentum (Isınma)" and "Isınmaya" in str(cell.value): cell.fill = fills["heat"]
            if col_name == "Regresyon (89)":
                val = str(cell.value)
                if val == "Güçlü Pozitif": cell.fill, cell.font = fills["pos_strong"], Font(color="FFFFFF", bold=True)
                elif val == "Güçlü Negatif": cell.fill, cell.font = fills["neg_strong"], Font(color="FFFFFF", bold=True)

            if col_name == "Analiz Notu":
                val = str(cell.value)
                if any(x in val for x in ["Sert Dağıtım", "Sert Satış"]): cell.fill, cell.font = fills["neg_strong"], Font(color="FFFFFF", bold=True)
                elif "Negatif" in val or any(x in val for x in ["Zayıf", "Tuzak", "Kontrollü Çıkış"]): cell.fill = fills["neg_light"]
                elif any(x in val for x in ["Agresif", "Güçlü"]): cell.fill, cell.font = fills["pos_strong"], Font(color="FFFFFF", bold=True)
                elif any(x in val for x in ["Sessiz Biriktirme", "Pozitif"]): cell.fill = fills["pos_light"]
                elif "Nötr" in val: cell.fill = fills["neutral"]

            if col_name in TL_FORMAT_COLS: cell.number_format = '#,##0.00 "₺"'
            elif col_name in PCT_FORMAT_COLS: cell.number_format = '0.00"%"'
            elif "Skoru" in str(col_name): cell.number_format = '0.00'

    for col_cells in ws.columns:
        column, col_name = col_cells[0].column_letter, col_cells[0].value
        if col_name == "Grafik": ws.column_dimensions[column].width = 20
        else:
            max_len = max([len(str(cell.value)) for cell in col_cells if cell.value] + [0])
            ws.column_dimensions[column].width = max_len + 3
    ws.freeze_panes = "A2"

def apply_conditional_formatting(ws):
    headers = {cell.value: cell.column_letter for cell in ws[1]}
    pos_f, neg_f = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid"), PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
    for col in ["Net Para Akışı (10G)", "Para Akış Oranı", "Akış Gücü Skoru", "10 Günlük Getiri (%)", "Toplam Net Para Akışı", "Ağırlıklı Akış Skoru", "Hacim Ağırlıklı Getiri (%)"]:
        if col in headers:
            rng = f"{headers[col]}2:{headers[col]}{ws.max_row}"
            ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], fill=pos_f))
            ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], fill=neg_f))

# =========================================================
# MODÜL 4: EXCEL ORKESTRASYON
# =========================================================
def save_report(res, clean_symbols, skipped, chart_data_bytes, sector_summary):
    res_sorted = res.sort_values("Akış Gücü Skoru", ascending=False)
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        res_sorted.to_excel(writer, sheet_name=SHEET_MAIN, index=False)
        res_sorted.head(10).to_excel(writer, sheet_name=SHEET_TOP, index=False)
        sector_summary.to_excel(writer, sheet_name=SHEET_SECTOR, index=True)
        pd.DataFrame({"Hisse Kodu": clean_symbols}).to_excel(writer, sheet_name=SHEET_SYMBOLS, index=False)
        if skipped: pd.DataFrame(skipped).to_excel(writer, sheet_name=SHEET_SKIPPED, index=False)
        
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            apply_excel_styling(ws)
            apply_conditional_formatting(ws)
            
            if sheet_name == SHEET_SECTOR:
                pie = PieChart()
                pie.add_data(Reference(ws, min_col=3, min_row=1, max_row=ws.max_row), titles_from_data=True)
                pie.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
                pie.title = "Sektörel Hacim Dağılımı"
                pie.title.overlay = False 
                pie.height = 12 
                pie.width = 20  
                pie.legend.position = 'r' 
                
                # 2- GRAFİĞİ BİLGİ HÜCRELERİNİN ALTINA YERLEŞTİR
                ws.add_chart(pie, f"A{ws.max_row + 2}")
                
            if sheet_name in [SHEET_MAIN, SHEET_TOP]:
                headers = {cell.value: cell.column_letter for cell in ws[1]}
                sym_col_idx = list(headers.keys()).index("Hisse Kodu") + 1
                chart_col_let = headers.get("Grafik")
                if chart_col_let:
                    for r_idx in range(2, ws.max_row + 1):
                        sym = ws.cell(row=r_idx, column=sym_col_idx).value
                        if sym in chart_data_bytes:
                            ws.add_image(Image(io.BytesIO(chart_data_bytes[sym])), f"{chart_col_let}{r_idx}")

def run_full_process():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    try:
        data, clean_symbols, yahoo_symbols, sector_map = fetch_bist_data()
        res, skipped, chart_data_bytes, sector_summary = perform_analysis(data, yahoo_symbols, sector_map)
        save_report(res, clean_symbols, skipped, chart_data_bytes, sector_summary)
        print(f"\n[BAŞARILI] Rapor Hazır: {OUTPUT_FILE}")
    except Exception as e: print(f"\n[SİSTEM HATASI]: {e}")

if __name__ == "__main__": run_full_process()